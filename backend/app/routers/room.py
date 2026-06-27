from fastapi import APIRouter, HTTPException, status
from pydantic import BaseModel
from typing import List, Optional
import math
import io
import openpyxl
from fastapi.responses import StreamingResponse

from app.models.room import RoomModel
from app.models.student import StudentModel
from beanie.operators import In

router = APIRouter(prefix="/drives", tags=["Rooms Logistics"])

class CreateRoomRequest(BaseModel):
    name: str
    capacity: int
    purpose: Optional[str] = "General"
    block: Optional[str] = None
    floor: Optional[str] = None

class AllocateRoomsRequest(BaseModel):
    student_status_filter: str = "present"
    target_room_purpose: str = "General"

class ClearRoomsRequest(BaseModel):
    target_room_purpose: str

@router.post("/{drive_id}/rooms", status_code=status.HTTP_201_CREATED)
async def create_room(drive_id: str, payload: CreateRoomRequest):
    """Create a physical room for a drive"""
    room = RoomModel(
        drive_id=drive_id,
        name=payload.name,
        capacity=payload.capacity,
        purpose=payload.purpose or "General",
        block=payload.block,
        floor=payload.floor,
        current_occupancy=0,
        is_locked=False
    )
    await room.insert()
    return {**room.model_dump(), "id": str(room.id)}

@router.get("/{drive_id}/rooms")
async def list_rooms(drive_id: str):
    """List all rooms and their real-time occupancies for a drive"""
    rooms = await RoomModel.find(RoomModel.drive_id == drive_id).to_list()
    # Serialize ObjectId
    return [{**r.model_dump(), "id": str(r.id)} for r in rooms]

@router.delete("/{drive_id}/rooms/{room_id}")
async def delete_room(drive_id: str, room_id: str):
    """Delete a room"""
    room = await RoomModel.get(room_id)
    if not room or room.drive_id != drive_id:
        raise HTTPException(status_code=404, detail="Room not found")
    await room.delete()
    return {"message": "Room deleted successfully"}

from fastapi import UploadFile, File
from app.services.xlsx_parser import process_room_upload

@router.post("/{drive_id}/rooms/upload")
async def upload_rooms(drive_id: str, file: UploadFile = File(...)):
    """Upload XLSX file for bulk room creation"""
    filename = file.filename or ""
    if not filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload an Excel file.")

    contents = await file.read()
    result = await process_room_upload(contents, drive_id)
    return result

@router.post("/{drive_id}/allocate-rooms")
async def allocate_rooms(drive_id: str, payload: AllocateRoomsRequest):
    """
    Advanced Purpose-Aware Logistics Engine.
    Finds unassigned students by status and distributes them to rooms by purpose.
    """
    # 1. Get available rooms that match the target purpose
    rooms = await RoomModel.find(
        RoomModel.drive_id == drive_id,
        RoomModel.purpose == payload.target_room_purpose
    ).to_list()
    if not rooms:
        raise HTTPException(status_code=400, detail=f"No rooms configured for purpose: {payload.target_room_purpose}")

    # 2. Get unassigned students matching the target status
    unassigned_students = await StudentModel.find(
        StudentModel.drive_id == drive_id,
        StudentModel.status == payload.student_status_filter,
        StudentModel.current_room_id == None
    ).to_list()

    allocated_count = 0

    # 3. Distribute students
    for student in unassigned_students:
        available_room = None
        for room in rooms:
            if room.current_occupancy < room.capacity and not room.is_locked:
                available_room = room
                break
        
        if not available_room:
            break # No more capacity
            
        # Allocate
        student.current_room_id = str(available_room.id)
        await student.save()
        
        available_room.current_occupancy += 1
        allocated_count += 1

    # 4. Save updated room capacities
    for room in rooms:
        await room.save()

    return {
        "message": "Allocation complete",
        "allocated_count": allocated_count,
        "unassigned_remaining": len(unassigned_students) - allocated_count
    }

@router.post("/{drive_id}/clear-rooms")
async def clear_rooms(drive_id: str, payload: ClearRoomsRequest):
    """
    Clears all rooms of a specific purpose, un-assigning students from them.
    Useful when a round (like GD) finishes and rooms need to be freed for the next batch.
    """
    # 1. Find the target rooms
    rooms = await RoomModel.find(
        RoomModel.drive_id == drive_id,
        RoomModel.purpose == payload.target_room_purpose
    ).to_list()
    
    if not rooms:
        return {"message": "No rooms found for this purpose.", "cleared_rooms": 0}

    room_ids = [str(r.id) for r in rooms]

    # 2. Find students currently assigned to these rooms and un-assign them
    # Note: We don't change their status here, just their physical room assignment
    students = await StudentModel.find(
        StudentModel.drive_id == drive_id,
        In(StudentModel.current_room_id, room_ids)
    ).to_list()

    for student in students:
        student.current_room_id = None
        await student.save()

    # 3. Reset room occupancies to 0
    for room in rooms:
        room.current_occupancy = 0
        await room.save()

    return {"message": f"Successfully cleared {len(rooms)} {payload.target_room_purpose} rooms.", "cleared_rooms": len(rooms)}

@router.get("/{drive_id}/stats/god-view")
async def god_view_stats(drive_id: str):
    """Fetch total shortlisted, checked-in, and pending for the top row of God View"""
    total_shortlisted = await StudentModel.find(
        StudentModel.drive_id == drive_id,
        # We assume they are either registered or shortlisted.
        # If they check in, status becomes 'present'
    ).count()

    checked_in = await StudentModel.find(
        StudentModel.drive_id == drive_id,
        StudentModel.status == "present"
    ).count()

    # The actual numbers for the dashboard
    return {
        "total_shortlisted": total_shortlisted,
        "checked_in": checked_in,
        "pending_arrival": total_shortlisted - checked_in
    }

@router.get("/{drive_id}/rooms/export")
async def export_room_data(drive_id: str, export_type: str, room_id: Optional[str] = None):
    """
    Export student lists to Excel.
    export_type can be:
    - qr_checkin: all students who have scanned QR and checked in
    - room_wise: all students allocated, separated by room (sheets)
    - specific_room: all students allocated to a specific room
    """
    wb = openpyxl.Workbook()
    
    # helper function to write data
    def write_sheet(ws, title, students):
        # Sheet names cannot contain invalid characters and are limited to 31 chars
        safe_title = "".join([c for c in title if c.isalnum() or c in " -_"])[:31]
        ws.title = safe_title
        headers = ["Unique ID", "Full Name", "Email", "Phone", "Status", "Check-In Time"]
        ws.append(headers)
        for s in students:
            check_in_str = s.check_in_time.strftime("%Y-%m-%d %H:%M:%S") if s.check_in_time else "N/A"
            ws.append([s.unique_id, s.full_name, s.email, s.phone, s.status, check_in_str])

    if export_type == "qr_checkin":
        # Get students who have checked in (have check_in_time)
        students = await StudentModel.find(
            StudentModel.drive_id == drive_id,
            StudentModel.check_in_time != None
        ).to_list()
        
        ws = wb.active
        write_sheet(ws, "QR Check-ins", students)
        filename = f"QR_Checkins.xlsx"
        
    elif export_type == "room_wise":
        # Get all rooms
        rooms = await RoomModel.find(RoomModel.drive_id == drive_id).to_list()
        
        # Get all allocated students
        students = await StudentModel.find(
            StudentModel.drive_id == drive_id,
            StudentModel.current_room_id != None
        ).to_list()
        
        # Group by room
        wb.remove(wb.active) # remove default sheet
        
        room_map = {str(r.id): r.name for r in rooms}
        
        grouped_students = {}
        for s in students:
            rid = str(s.current_room_id)
            if rid not in grouped_students:
                grouped_students[rid] = []
            grouped_students[rid].append(s)
            
        for rid, room_students in grouped_students.items():
            room_name = room_map.get(rid, f"Room_{rid}")
            ws = wb.create_sheet(title=room_name)
            write_sheet(ws, room_name, room_students)
            
        if not wb.sheetnames:
            wb.create_sheet("Empty")
            
        filename = f"Room_Wise_Allocation.xlsx"

    elif export_type == "specific_room":
        if not room_id:
            raise HTTPException(status_code=400, detail="room_id is required for specific_room export")
            
        room = await RoomModel.get(room_id)
        if not room:
            raise HTTPException(status_code=404, detail="Room not found")
            
        students = await StudentModel.find(
            StudentModel.drive_id == drive_id,
            StudentModel.current_room_id == room_id
        ).to_list()
        
        ws = wb.active
        write_sheet(ws, room.name, students)
        filename = f"{room.name}_Allocation.xlsx"
    else:
        raise HTTPException(status_code=400, detail="Invalid export_type")

    # Save to stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        output,
        headers=headers,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

